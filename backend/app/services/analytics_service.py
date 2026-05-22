"""
Analytics and Reporting service - Advanced dashboard data and exports
"""
from datetime import datetime, timedelta
from typing import Dict, List, Optional
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, and_, or_, extract
from uuid import UUID
import csv
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from app.models import Job, Candidate, Application, User, Agency, JobStatus, ApplicationStatus, CandidateStatus, ClientCompany


class AnalyticsService:
    """Analytics and reporting service"""
    
    # ============= Dashboard Analytics =============
    
    @staticmethod
    async def get_dashboard_overview(
        db: AsyncSession,
        agency_id: UUID,
        period_days: int = 30
    ) -> Dict:
        """
        Get high-level dashboard metrics
        
        Args:
            db: Database session
            agency_id: Agency ID
            period_days: Period for metrics (default 30 days)
            
        Returns:
            Dashboard overview metrics
        """
        period_start = datetime.utcnow() - timedelta(days=period_days)
        
        # Total jobs
        total_jobs_result = await db.execute(
            select(func.count(Job.id)).where(Job.agency_id == agency_id)
        )
        total_jobs = total_jobs_result.scalar()
        
        # Active jobs
        active_jobs_result = await db.execute(
            select(func.count(Job.id)).where(
                Job.agency_id == agency_id,
                Job.status == JobStatus.active
            )
        )
        active_jobs = active_jobs_result.scalar()
        
        # Total candidates
        total_candidates_result = await db.execute(
            select(func.count(Candidate.id)).where(Candidate.agency_id == agency_id)
        )
        total_candidates = total_candidates_result.scalar()
        
        # Active candidates
        active_candidates_result = await db.execute(
            select(func.count(Candidate.id)).where(
                Candidate.agency_id == agency_id,
                Candidate.status == CandidateStatus.active
            )
        )
        active_candidates = active_candidates_result.scalar()
        
        # Total applications (this period)
        applications_result = await db.execute(
            select(func.count(Application.id)).where(
                Application.agency_id == agency_id,
                Application.created_at >= period_start
            )
        )
        applications_this_period = applications_result.scalar()

        # Total clients
        clients_result = await db.execute(
            select(func.count(ClientCompany.id)).where(ClientCompany.agency_id == agency_id)
        )
        total_clients = clients_result.scalar()
        
        # Successful placements (this period)
        placements_result = await db.execute(
            select(func.count(Application.id)).where(
                Application.agency_id == agency_id,
                Application.status == ApplicationStatus.hired,
                Application.created_at >= period_start
            )
)
        placements = placements_result.scalar()
        
        # Average time to hire (days)
        time_to_hire_result = await db.execute(
            select(
                func.avg(
                    extract('epoch', Application.updated_at - Application.created_at) / 86400
                )
            ).where(
                Application.agency_id == agency_id,
                Application.status == ApplicationStatus.hired,
                Application.created_at >= period_start
            )
        )
        avg_time_to_hire = time_to_hire_result.scalar() or 0
        
        # Success rate
        success_rate = (placements / applications_this_period * 100) if applications_this_period > 0 else 0
        
        return {
            "period_days": period_days,
            "total_jobs": total_jobs,
            "active_jobs": active_jobs,
            "total_candidates": total_candidates,
            "active_candidates": active_candidates,
            "applications_this_period": applications_this_period,
            "total_clients": total_clients,
            "placements": placements,
            "avg_time_to_hire_days": round(avg_time_to_hire, 1),
            "success_rate_percentage": round(success_rate, 2)
        }
    
    @staticmethod
    async def get_applications_over_time(
        db: AsyncSession,
        agency_id: UUID,
        period_days: int = 30
    ) -> List[Dict]:
        """Get applications count over time (for chart)"""
        period_start = datetime.utcnow() - timedelta(days=period_days)
        
        result = await db.execute(
            select(
                func.date(Application.created_at).label('date'),
                func.count(Application.id).label('count')
            ).where(
                Application.agency_id == agency_id,
                Application.created_at >= period_start
            ).group_by(
                func.date(Application.created_at)
            ).order_by(
                func.date(Application.created_at)
            )
        )
        
        data = result.all()
        
        return [
            {
                "date": row.date.isoformat(),
                "count": row.count
            }
            for row in data
        ]
    
    @staticmethod
    async def get_applications_by_status(
        db: AsyncSession,
        agency_id: UUID
    ) -> List[Dict]:
        """Get applications distribution by status (for pie chart)"""
        result = await db.execute(
            select(
                Application.status,
                func.count(Application.id).label('count')
            ).where(
                Application.agency_id == agency_id
            ).group_by(
                Application.status
            )
        )
        
        data = result.all()
        
        return [
            {
                "status": row.status.value,
                "count": row.count
            }
            for row in data
        ]
    
    @staticmethod
    async def get_top_performing_jobs(
        db: AsyncSession,
        agency_id: UUID,
        limit: int = 10
    ) -> List[Dict]:
        """Get top performing jobs by applications count"""
        result = await db.execute(
            select(
                Job.id,
                Job.title,
                Job.reference,
                func.count(Application.id).label('applications_count'),
                func.count(
                    func.nullif(Application.status == ApplicationStatus.hired, False)
                ).label('hires_count')
            ).join(
                Application, Application.job_id == Job.id
            ).where(
                Job.agency_id == agency_id
            ).group_by(
                Job.id, Job.title, Job.reference
            ).order_by(
                func.count(Application.id).desc()
            ).limit(limit)
        )
        
        data = result.all()
        
        return [
            {
                "job_id": str(row.id),
                "title": row.title,
                "reference": row.reference,
                "applications_count": row.applications_count,
                "hires_count": row.hires_count,
                "conversion_rate": round((row.hires_count / row.applications_count * 100) if row.applications_count > 0 else 0, 2)
            }
            for row in data
        ]
    
    @staticmethod
    async def get_recruiter_performance(
        db: AsyncSession,
        agency_id: UUID
    ) -> List[Dict]:
        """Get recruiter performance metrics"""
        result = await db.execute(
            select(
                User.id,
                User.first_name,
                User.last_name,
                func.count(Application.id).label('total_applications'),
                func.count(
                    func.nullif(Application.status == ApplicationStatus.hired, False)
                ).label('successful_hires')
            ).join(
                Application, Application.assigned_to == User.id
            ).where(
                User.agency_id == agency_id
            ).group_by(
                User.id, User.first_name, User.last_name
            ).order_by(
                func.count(
                    func.nullif(Application.status == ApplicationStatus.hired, False)
                ).desc()
            )
        )
        
        data = result.all()
        
        return [
            {
                "recruiter_id": str(row.id),
                "name": f"{row.first_name} {row.last_name}",
                "total_applications": row.total_applications,
                "successful_hires": row.successful_hires,
                "success_rate": round((row.successful_hires / row.total_applications * 100) if row.total_applications > 0 else 0, 2)
            }
            for row in data
        ]
    
    @staticmethod
    async def get_source_effectiveness(
        db: AsyncSession,
        agency_id: UUID
    ) -> List[Dict]:
        """Get candidate source effectiveness"""
        result = await db.execute(
            select(
                Candidate.source,
                func.count(Candidate.id).label('total_candidates'),
                func.sum(Candidate.applications_count).label('total_applications'),
                func.sum(Candidate.placements_count).label('total_placements')
            ).where(
                Candidate.agency_id == agency_id
            ).group_by(
                Candidate.source
            ).order_by(
                func.sum(Candidate.placements_count).desc()
            )
        )
        
        data = result.all()
        
        return [
            {
                "source": row.source.value,
                "candidates_count": row.total_candidates,
                "applications_count": row.total_applications or 0,
                "placements_count": row.total_placements or 0,
                "conversion_rate": round(((row.total_placements or 0) / row.total_candidates * 100) if row.total_candidates > 0 else 0, 2)
            }
            for row in data
        ]
    
    # ============= Export Functions =============
    
    @staticmethod
    async def export_applications_to_csv(
        db: AsyncSession,
        agency_id: UUID,
        filters: Optional[Dict] = None
    ) -> str:
        """Export applications to CSV"""
        # Get applications
        query = select(
            Application,
            Job.title.label('job_title'),
            Job.reference.label('job_reference'),
            Candidate.first_name,
            Candidate.last_name,
            Candidate.email
        ).join(
            Job, Application.job_id == Job.id
        ).join(
            Candidate, Application.candidate_id == Candidate.id
        ).where(
            Application.agency_id == agency_id
        )
        
        # Apply filters if provided
        if filters:
            if filters.get('status'):
                query = query.where(Application.status == filters['status'])
            if filters.get('created_after'):
                query = query.where(Application.created_at >= filters['created_after'])
        
        result = await db.execute(query)
        data = result.all()
        
        # Create CSV
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Header
        writer.writerow([
            'Application ID', 'Job Title', 'Job Reference', 'Candidate Name',
            'Candidate Email', 'Status', 'Match Score', 'Created At', 'Updated At'
        ])
        
        # Data rows
        for row in data:
            writer.writerow([
                str(row.Application.id),
                row.job_title,
                row.job_reference,
                f"{row.first_name} {row.last_name}",
                row.email,
                row.Application.status.value,
                row.Application.match_score or 'N/A',
                row.Application.created_at.isoformat(),
                row.Application.updated_at.isoformat()
            ])
        
        return output.getvalue()
    
    @staticmethod
    async def export_candidates_to_excel(
        db: AsyncSession,
        agency_id: UUID
    ) -> bytes:
        """Export candidates to Excel with formatting"""
        # Get candidates
        result = await db.execute(
            select(Candidate).where(Candidate.agency_id == agency_id)
        )
        candidates = result.scalars().all()
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Candidates"
        
        # Header style
        header_fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        # Headers
        headers = [
            'Name', 'Email', 'Phone', 'Location', 'Experience (Years)',
            'Skills', 'Education', 'Status', 'Applications', 'Placements'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Data rows
        for row_idx, candidate in enumerate(candidates, 2):
            ws.cell(row=row_idx, column=1, value=candidate.full_name)
            ws.cell(row=row_idx, column=2, value=candidate.email)
            ws.cell(row=row_idx, column=3, value=candidate.phone or '')
            ws.cell(row=row_idx, column=4, value=f"{candidate.city or ''}, {candidate.province or ''}")
            ws.cell(row=row_idx, column=5, value=candidate.years_of_experience)
            ws.cell(row=row_idx, column=6, value=', '.join(candidate.skills) if candidate.skills else '')
            ws.cell(row=row_idx, column=7, value=candidate.education_level or '')
            ws.cell(row=row_idx, column=8, value=candidate.status.value)
            ws.cell(row=row_idx, column=9, value=candidate.applications_count)
            ws.cell(row=row_idx, column=10, value=candidate.placements_count)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()


    # ============= Custom Reports & Predictions =============

    @staticmethod
    async def get_custom_report(
        db: AsyncSession,
        agency_id: UUID,
        metrics: List[str],
        group_by: str = "month",
        period_days: int = 90,
        company_id: Optional[UUID] = None,
    ) -> Dict:
        """Dynamic custom report builder. Supported metrics: applications, hires, rejections, interviews, offers, time_to_hire, source_breakdown."""
        period_start = datetime.utcnow() - timedelta(days=period_days)
        base_filters = [Application.agency_id == agency_id, Application.created_at >= period_start]
        if company_id:
            base_filters.append(Application.client_company_id == company_id)

        results = {}

        if "applications" in metrics:
            r = await db.execute(select(func.count(Application.id)).where(*base_filters))
            results["applications"] = r.scalar() or 0

        if "hires" in metrics:
            r = await db.execute(select(func.count(Application.id)).where(*base_filters, Application.status == ApplicationStatus.hired))
            results["hires"] = r.scalar() or 0

        if "rejections" in metrics:
            r = await db.execute(select(func.count(Application.id)).where(*base_filters, Application.status == ApplicationStatus.rejected))
            results["rejections"] = r.scalar() or 0

        if "interviews" in metrics:
            r = await db.execute(select(func.count(Application.id)).where(*base_filters, Application.status.in_([ApplicationStatus.interview_scheduled, ApplicationStatus.interviewed])))
            results["interviews"] = r.scalar() or 0

        if "offers" in metrics:
            r = await db.execute(select(func.count(Application.id)).where(*base_filters, Application.status.in_([ApplicationStatus.offer_made, ApplicationStatus.offer_accepted])))
            results["offers"] = r.scalar() or 0

        if "source_breakdown" in metrics:
            r = await db.execute(
                select(Application.source, func.count(Application.id)).where(*base_filters).group_by(Application.source)
            )
            results["source_breakdown"] = {row[0].value: row[1] for row in r.all()}

        # Group by month if requested
        if group_by == "month":
            r = await db.execute(
                select(
                    func.date_trunc('month', Application.created_at).label("month"),
                    func.count(Application.id).label("count"),
                ).where(*base_filters).group_by(func.date_trunc('month', Application.created_at)).order_by("month")
            )
            results["by_month"] = [
                {"month": row.month.strftime("%Y-%m"), "count": row.count}
                for row in r.all()
            ]

        return {
            "period_days": period_days,
            "metrics": metrics,
            "group_by": group_by,
            "data": results,
        }

    @staticmethod
    async def get_predictions(
        db: AsyncSession,
        agency_id: UUID,
        company_id: Optional[UUID] = None,
    ) -> Dict:
        """AI-powered predictive analytics using historical hiring data."""
        import json, os
        import openai

        # Gather historical data
        period_start = datetime.utcnow() - timedelta(days=180)
        base = [Application.agency_id == agency_id, Application.created_at >= period_start]
        if company_id:
            base.append(Application.client_company_id == company_id)

        total_r = await db.execute(select(func.count(Application.id)).where(*base))
        hired_r = await db.execute(select(func.count(Application.id)).where(*base, Application.status == ApplicationStatus.hired))
        rejected_r = await db.execute(select(func.count(Application.id)).where(*base, Application.status == ApplicationStatus.rejected))

        total = total_r.scalar() or 0
        hired = hired_r.scalar() or 0
        rejected = rejected_r.scalar() or 0
        hire_rate = round((hired / total) * 100, 1) if total > 0 else 0

        source_r = await db.execute(
            select(Application.source, func.count(Application.id).label("cnt"))
            .where(*base).group_by(Application.source).order_by(func.count(Application.id).desc()).limit(1)
        )
        top_source_row = source_r.first()
        top_source = top_source_row[0].value if top_source_row else "direct"

        api_key = os.getenv("OPENAI_API_KEY", "")
        if not api_key:
            return {
                "hire_rate": hire_rate,
                "top_source": top_source,
                "predictions": {"message": "AI predictions require OpenAI API key"},
            }

        client = openai.AsyncOpenAI(api_key=api_key)
        prompt = f"""You are a South African recruitment analytics expert. Based on this hiring data from the last 6 months, provide actionable predictions:

Data:
- Total applications: {total}
- Hired: {hired} ({hire_rate}% hire rate)
- Rejected: {rejected}
- Top performing source: {top_source}

Return JSON with:
{{
  "estimated_time_to_fill_days": <integer>,
  "best_candidate_source": "{top_source}",
  "salary_competitiveness": "competitive | below_market | above_market",
  "hiring_velocity_trend": "increasing | stable | decreasing",
  "recommendations": ["recommendation 1", "recommendation 2", "recommendation 3"],
  "risk_factors": ["risk 1", "risk 2"]
}}
"""
        response = await client.chat.completions.create(
            model=os.getenv("OPENAI_MODEL", "gpt-4"),
            messages=[{"role": "user", "content": prompt}],
            response_format={"type": "json_object"},
            temperature=0.3,
        )
        predictions = json.loads(response.choices[0].message.content)
        return {
            "hire_rate": hire_rate,
            "top_source": top_source,
            "total_applications_6mo": total,
            "predictions": predictions,
        }


# Create service instance
analytics_service = AnalyticsService()
